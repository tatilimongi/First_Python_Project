from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Ler pasta de trabalho e planilha
wb = load_workbook('data/barchart.xlsx')
sheet = wb['Relatorio']

# Referencias das linhas e colunas
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# Incluir formulas manualmente
# sheet['B6'] = "=SUM(B2:B5)"
# sheet['B6'].style = 'Currency'

# Incluir formulas dinamicamente

# A partir da coluna B, depois dos títulos
for i in range(min_column + 1, max_column + 1): 
    letter = get_column_letter(i)
    # Exibir resultados depois da ultima linha / somar a partir da 2a linha
    sheet[f'{letter}{max_row + 1}'] = f'=SUM({letter}{min_row + 1}:{letter}{max_row})'
    sheet[f'{letter}{max_row + 1}'].style = 'Currency'

wb.save('teste.xlsx')